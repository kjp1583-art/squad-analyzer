#!/usr/bin/env python3
"""십이귀월(상현/하현) 로스터를 시트 스냅샷에서 계산한다 — index.html computeAssessments의 파이썬 이식.

분석기 호스트가 꺼져 있어도 클라우드에서 명단을 뽑고 웹훅을 쏠 수 있게 한다.
산식·가중치·정렬·리그 분할은 웹과 동일하게 유지할 것 (한쪽만 고치면 명단이 갈라진다).
"""
import argparse, json, math, os, re, sys, urllib.request
from collections import defaultdict

import openpyxl

TIER_MIN_GAMES, TIER_MIN_EVAL = 10, 5
SHRINK_WR, SHRINK_AI, SHRINK_EVAL, SHRINK_SOLOWR = 10, 10, 5, 20
W_SOLO, W_AI, W_WR, W_SOLOWR, W_MVP, W_TROLL = 0.35, 0.20, 0.15, 0.10, 0.10, 0.10
RECENT_DAYS, RECENT_MIN = 30, 5


def norm(s):
    return re.sub(r"\s+", "", str(s or "").lower())


def tnorm(s):
    return norm(str(s or "").split("#")[0])


def mean(xs):
    xs = [x for x in xs if x is not None]
    return sum(xs) / len(xs) if xs else None


def std(xs, mu):
    xs = [x for x in xs if x is not None]
    if len(xs) < 2 or mu is None: return None
    return math.sqrt(sum((x - mu) ** 2 for x in xs) / (len(xs) - 1))


def read_tab(wb, name):
    if name not in wb.sheetnames: return [], []
    it = wb[name].iter_rows(values_only=True)
    try: head = [str(c or "").strip() for c in next(it)]
    except StopIteration: return [], []
    return head, [list(r) for r in it]


def load(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    h, rows = read_tab(wb, "CLASSIC_NORMAL")
    idx = {k: h.index(k) for k in h}
    raw = [{k: (r[i] if i < len(r) else None) for k, i in idx.items()} for r in rows]
    raw = [r for r in raw if r.get("소환사명")]

    # 내부티어
    _h, trows = read_tab(wb, "CLAN_TIERS")
    tier_of_raw = {tnorm(r[0]): str(r[1]).strip() for r in trows if r and r[0] and len(r) > 1 and r[1]}

    # 부계정 통합
    _h, lrows = read_tab(wb, "LINK_ACCOUNT")
    alt_to_main = {tnorm(r[1]): str(r[0]).strip() for r in lrows if r and len(r) > 1 and r[0] and r[1]}

    # 솔랭 + 과거 3시즌 최고 블렌드
    h, srows = read_tab(wb, "SOLO_RANK")
    def col(hh, *names):
        for n in names:
            if n in hh: return hh.index(n)
        return -1
    ni, si, wi, li = col(h, "닉네임"), col(h, "점수"), col(h, "솔랭승"), col(h, "솔랭패")
    solo = {}
    for r in srows:
        if ni < 0 or si < 0 or len(r) <= si or not r[ni]: continue
        try: sc = float(r[si])
        except (TypeError, ValueError): continue
        try: w, l = int(r[wi]), int(r[li])
        except (TypeError, ValueError, IndexError): w = l = 0
        solo[tnorm(r[ni])] = {"score": sc, "wins": w, "losses": l,
                              "wr": (w / (w + l) * 100) if (w + l) else None, "cur": sc}
    h, prows = read_tab(wb, "PEAK_SEASONS")
    pn, pp = col(h, "닉네임"), col(h, "점수")
    if pn >= 0 and pp >= 0:
        for r in prows:
            if len(r) <= max(pn, pp) or not r[pn]: continue
            try: pk = float(r[pp])
            except (TypeError, ValueError): continue
            k = tnorm(r[pn])
            if k in solo: solo[k]["score"] = (solo[k]["score"] + pk) / 2.0
            else: solo[k] = {"score": pk, "wins": 0, "losses": 0, "wr": None, "cur": None}
    return raw, tier_of_raw, alt_to_main, solo


def build_identity(raw, alt_to_main):
    """PUUID 최신닉 → 대표닉, 부계정은 본계정으로 흡수 (웹 buildIdentity와 동일)."""
    latest = {}
    for i, r in enumerate(raw):
        pu, nm, d = r.get("PUUID"), r.get("소환사명"), str(r.get("날짜") or "")
        if not pu or not nm: continue
        cur = latest.get(pu)
        if not cur or d > cur[1] or (d == cur[1] and i >= cur[2]): latest[pu] = (nm, d, i)
    canon_by_puuid = {pu: v[0] for pu, v in latest.items()}

    name_to_canon, canon_aliases = {}, defaultdict(set)
    for r in raw:
        pu, nm = r.get("PUUID"), r.get("소환사명")
        if not nm: continue
        canon = canon_by_puuid.get(pu, nm) if pu else nm
        k = norm(nm)
        if pu and pu in canon_by_puuid: name_to_canon[k] = canon
        elif k not in name_to_canon: name_to_canon[k] = nm
        canon_aliases[canon].add(nm)
    tagless = {tnorm(c): c for c in canon_aliases}

    def resolve_alt(nm):
        main = alt_to_main.get(tnorm(nm))
        if not main: return nm
        return name_to_canon.get(norm(main)) or tagless.get(tnorm(main)) or main

    def canon_of_row(r):
        pu = r.get("PUUID")
        if pu and pu in canon_by_puuid: return resolve_alt(canon_by_puuid[pu])
        return resolve_alt(r.get("소환사명") or "")

    return canon_of_row, canon_aliases, resolve_alt


def compute(path, today=None):
    raw, tier_of_raw, alt_to_main, solo = load(path)
    canon_of_row, canon_aliases, resolve_alt = build_identity(raw, alt_to_main)

    # 게임 단위 중복 제거 후 대표닉으로 묶기
    seen, games = set(), defaultdict(list)
    for r in raw:
        gid = r.get("게임ID")
        if not gid: continue
        key = (gid, r.get("PUUID") or r.get("소환사명"), r.get("챔피언"), r.get("포지션"), r.get("진영"))
        if key in seen: continue
        seen.add(key)
        games[gid].append(r)
    players = defaultdict(list)
    for rows in games.values():
        for r in rows: players[canon_of_row(r)].append(r)

    eval_gids = {gid for gid, rows in games.items()
                 if any(str(r.get("매치평가") or "") in ("MVP", "역적", "ACE") for r in rows)}

    def tier_of(name):
        t = tier_of_raw.get(tnorm(name))
        if t: return t
        for a in canon_aliases.get(name, ()):
            t = tier_of_raw.get(tnorm(a))
            if t: return t
        return None

    import datetime
    cut = (today or datetime.date.today()) - datetime.timedelta(days=RECENT_DAYS)

    members = []
    for name, parts in players.items():
        t = tier_of(name)
        if not t: continue
        w = l = mvp = ace = troll = 0
        ai_sum = 0.0; ai_n = 0; eg = set(); recent = 0
        for r in parts:
            res = str(r.get("결과") or "")
            if res == "승리": w += 1
            elif res == "패배": l += 1
            e = str(r.get("매치평가") or "")
            if e == "MVP": mvp += 1
            elif e == "ACE": ace += 1
            elif e == "역적": troll += 1
            if r.get("게임ID") in eval_gids: eg.add(r["게임ID"])
            try: ai_sum += float(r.get("점수")); ai_n += 1
            except (TypeError, ValueError): pass
            if res in ("승리", "패배"):
                m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", str(r.get("날짜") or ""))
                if m and datetime.date(int(m[1]), int(m[2]), int(m[3])) >= cut: recent += 1
        g, eval_g = w + l, len(eg)
        cands = [solo.get(tnorm(name))] + [solo.get(tnorm(a)) for a in canon_aliases.get(name, ())]
        cands = [c for c in cands if c]
        sr = next((c for c in cands if c.get("wr") is not None), cands[0] if cands else None)
        members.append({"name": name, "tier": t, "recent": recent, "g": g, "w": w, "evalG": eval_g,
                        "mvp": mvp, "ace": ace, "troll": troll, "aiSum": ai_sum, "aiN": ai_n,
                        "wr": (w / g * 100) if g else None,
                        "mvpRate": ((mvp + 0.5 * ace) / eval_g * 100) if eval_g else None,
                        "trollRate": (troll / eval_g * 100) if eval_g else None,
                        "avgAI": (ai_sum / ai_n) if ai_n else None,
                        "soloScore": sr["score"] if sr else None,
                        "soloWR": sr["wr"] if sr else None,
                        "soloW": sr["wins"] if sr else 0, "soloL": sr["losses"] if sr else 0})

    elig = [m for m in members if m["g"] >= TIER_MIN_GAMES]
    if not elig: return []
    g_wr = mean([m["wr"] for m in elig]) or 50
    g_ai = mean([m["avgAI"] for m in elig if m["avgAI"] is not None]) or 0
    g_mvp = mean([m["mvpRate"] for m in elig if m["evalG"] >= TIER_MIN_EVAL]) or 10
    g_tr = mean([m["trollRate"] for m in elig if m["evalG"] >= TIER_MIN_EVAL]) or 10
    g_swr = mean([m["soloWR"] for m in elig if m["soloWR"] is not None]) or 50

    for m in elig:
        sh = {"wr": (m["w"] + SHRINK_WR * (g_wr / 100)) / (m["g"] + SHRINK_WR) * 100,
              "ai": ((m["aiSum"] + SHRINK_AI * g_ai) / (m["aiN"] + SHRINK_AI)) if m["aiN"] > 0 else None,
              "mvp": None, "tr": None, "solo": m["soloScore"], "solowr": None}
        if m["evalG"] >= TIER_MIN_EVAL:
            sh["mvp"] = ((m["mvp"] + 0.5 * m["ace"]) + SHRINK_EVAL * (g_mvp / 100)) / (m["evalG"] + SHRINK_EVAL) * 100
            sh["tr"] = (m["troll"] + SHRINK_EVAL * (g_tr / 100)) / (m["evalG"] + SHRINK_EVAL) * 100
        if m["soloWR"] is not None:
            sh["solowr"] = (m["soloW"] + SHRINK_SOLOWR * (g_swr / 100)) / (m["soloW"] + m["soloL"] + SHRINK_SOLOWR) * 100
        m["sh"] = sh

    keys = ("wr", "ai", "mvp", "tr", "solo", "solowr")
    gm = {k: mean([m["sh"][k] for m in elig]) for k in keys}
    gs = {k: (std([m["sh"][k] for m in elig], gm[k]) or 1) for k in keys}

    def z(sh, k):
        return None if sh[k] is None or gm[k] is None else (sh[k] - gm[k]) / gs[k]

    powered = []
    for m in elig:
        if m["sh"]["solo"] is None or m["recent"] < RECENT_MIN: continue
        terms = []
        for wgt, k, sign in ((W_SOLO, "solo", 1), (W_AI, "ai", 1), (W_WR, "wr", 1),
                             (W_SOLOWR, "solowr", 1), (W_MVP, "mvp", 1), (W_TROLL, "tr", -1)):
            zz = z(m["sh"], k)
            if zz is not None: terms.append((wgt, sign * zz))
        ws = sum(t[0] for t in terms) or 1.0
        powered.append({"name": m["name"], "tier": m["tier"],
                        "power": sum(w_ * z_ for w_, z_ in terms) / ws,
                        "solo": m["sh"]["solo"] if m["sh"]["solo"] is not None else -1e9})
    powered.sort(key=lambda p: (-p["power"], -p["solo"], p["name"]))

    # [2026-08-07 사장님 지시 — 동서 구분 삭제] 상현 6 = 0·1티어 상위 6 · 하현 6 = 2·3티어 상위 6
    out = []
    west = [p for p in powered if str(p["tier"])[:1] in ("0", "1")]
    east = [p for p in powered if str(p["tier"])[:1] not in ("0", "1")]
    for i, p in enumerate(west[:6]): out.append({**p, "rank": "상현", "title": f"상현 {i+1}"})
    for i, p in enumerate(east[:6]): out.append({**p, "rank": "하현", "title": f"하현 {i+1}"})
    return out


def post(roster, url):
    sangs = [r for r in roster if r["rank"] == "상현"]
    hahas = [r for r in roster if r["rank"] == "하현"]
    sang = "\n".join(f"`{i+1}` **{m['name'].split('#')[0]}**" for i, m in enumerate(sangs)) or "—"
    haha = "\n".join(f"`{i+1}` **{m['name'].split('#')[0]}**" for i, m in enumerate(hahas)) or "—"
    embeds = [{"title": "⚔️　십이귀월 ( 十二鬼月 )　현황　⚔️", "color": 0x9B1B1B,
               "fields": [{"name": "🗡　상현 ( 上弦 )", "value": sang, "inline": True},
                          {"name": "🌙　하현 ( 下弦 )", "value": haha, "inline": True}],
               "footer": {"text": "스쿼드해체분석기 · squad.gg — 상현=0·1티어 상위 6 · 하현=2·3티어 상위 6"}}]
    payload = {"content": "🩸🩸　**십이귀월 현재 명단**　🩸🩸", "embeds": embeds}
    req = urllib.request.Request(url, data=json.dumps(payload).encode("utf-8"),
                                 headers={"Content-Type": "application/json", "User-Agent": "squad-ci"})
    with urllib.request.urlopen(req, timeout=20) as r:
        print("발송 완료:", r.status)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--post", action="store_true", help="웹훅 발송(미지정 시 출력만)")
    a = ap.parse_args()
    roster = compute(a.xlsx)
    for r in roster:
        print(f"{r['title']:<12} {r['name']:<22} power={r['power']:+.3f} solo={r['solo']:.0f} tier={r['tier']}")
    if a.post:
        raw = os.environ.get("APP_SECRETS_B64", "")
        import base64
        src = base64.b64decode(re.sub(r"\s", "", raw)).decode("utf-8", "replace")
        m = re.search(r'^\s*RESULT_WEBHOOK_URL\s*=\s*["\']([^"\']+)["\']', src, re.M)
        if not m: sys.exit("RESULT_WEBHOOK_URL 없음")
        post(roster, m.group(1))


if __name__ == "__main__":
    main()
