# 주간 AI 주간평 다이제스트 생성기 (2026-08-10 상시화 — 사장님 지시: 신규 지표 반영)
#
# 사용: python3 tooling/weekly_digest.py <이번주 월요일 YYYY-MM-DD>  (예: 2026-08-10 실행 시 인자 2026-08-03)
#   → digest_<주차>.json 생성. 통산 10판↑ 전원, 이번주(월~일)/전주 구분.
#
# 포함 지표:
#   기본: KDA·AI점수·매치평가(MVP/ACE/역적)·분당딜·분당CS·킬관여·시야·솔킬·직전평가(ai_eval_latest.json)
#   신규(2026-08-10 추가, 웹과 같은 산식):
#     🍀 팀운: 판마다 [팀원(본인 제외) 통산AI 평균 − 상대 평균]을 '매칭 상정 실력(5f−4m)'으로 회귀한 잔차 백분위
#     ⚖️ 짊어진무게: 매칭 상정 실력(5f−4m) 평균 백분위 — 매칭이 매긴 몸값
#     🔀 부포지션: 주포지션 외 라인들의 승률·AI를 그 포지션 5판↑ 비교군 백분위로, 판수 가중(3판↑만)
#     😤 억울지수: GRUDGE 탭 값 그대로(맞라인 기대우위율−실제우위율, +면 티어 과대평가 주장에 근거)
import openpyxl, json, re, sys, bisect, datetime as _dt
from collections import defaultdict

def kda_str(v):
    # 🩹 [2026-08-17] USER_ENTERED 백필이 "9/4/3"을 날짜(2009-04-03)로 바꾼 셀 복원 — YY/M/D 역산
    if isinstance(v, (_dt.datetime, _dt.date)):
        return f'{v.year - 2000}/{v.month}/{v.day}'
    return str(v or '')

def metrics(s):
    o = {}
    for t in str(s or '').split('|'):
        m = re.match(r'^([a-z]+)(-?\d+(?:\.\d+)?)$', t.strip())
        if m: o[m.group(1)] = float(m.group(2))
    return o if o.get('m') else {}

def build(week_mon, xlsx='squad_sheet.xlsx', prev_path='ai_eval_latest.json'):
    import datetime
    w0 = datetime.date.fromisoformat(week_mon)
    w1 = w0 + datetime.timedelta(days=6)
    p0 = w0 - datetime.timedelta(days=7); p1 = w0 - datetime.timedelta(days=1)
    W0, W1, P0, P1 = str(w0), str(w1), str(p0), str(p1)
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    try: prev_eval = json.load(open(prev_path))['evals']
    except Exception: prev_eval = {}

    latest, rows_all = {}, []
    for tab in ('CLASSIC_NORMAL', 'KIWI_KIWI'):
        if tab not in wb.sheetnames: continue
        rows = list(wb[tab].values)
        if not rows: continue
        ix = {n: i for i, n in enumerate(rows[0]) if n}
        def gv(r, name):
            i = ix.get(name, -1)
            return r[i] if 0 <= i < len(r) else None
        for r in rows[1:]:
            pu, nm, d = str(gv(r, 'PUUID') or ''), str(gv(r, '소환사명') or ''), str(gv(r, '날짜') or '')
            rows_all.append((r, dict(ix)))
            # 대표닉: 태그 있는 이름만·챔피언명 유출 행 제외 (2026-08-10 카시오페아 사고 가드)
            if pu and nm and '#' in nm and str(gv(r, '챔피언') or '').replace(' ', '') != nm.replace(' ', ''):
                cur = latest.get(pu)
                if not cur or d >= cur[1]: latest[pu] = (nm, d)
    def canon(pu, nm):
        if pu in latest: return latest[pu][0]
        return nm if '#' in nm else ''
    def gvx(r, ix, name):
        i = ix.get(name, -1)
        return r[i] if 0 <= i < len(r) else None

    P = defaultdict(lambda: {'tot': {'g': 0, 'w': 0, 'mvp': 0, 'ace': 0, 'bad': 0,
                                     'champs': defaultdict(int), 'pos': defaultdict(int), 'aiS': 0.0, 'aiN': 0},
                             'tw': [], 'pw': {'g': 0, 'w': 0},
                             'posst': defaultdict(lambda: {'g': 0, 'w': 0, 'aiS': 0.0, 'aiN': 0})})
    games = defaultdict(list)   # (협곡만) 팀운·무게 계산용: gid -> [(canon, win, ai?)]
    for r, ix in rows_all:
        res = str(gvx(r, ix, '결과') or '')
        if res not in ('승리', '패배'): continue
        nm = canon(str(gvx(r, ix, 'PUUID') or ''), str(gvx(r, ix, '소환사명') or ''))
        if not nm: continue
        d = str(gvx(r, ix, '날짜') or '')[:10]
        e = P[nm]; t = e['tot']
        t['g'] += 1; t['w'] += (res == '승리')
        ev = str(gvx(r, ix, '매치평가') or '')
        if ev == 'MVP': t['mvp'] += 1
        elif ev == 'ACE': t['ace'] += 1
        elif ev == '역적': t['bad'] += 1
        ch = str(gvx(r, ix, '챔피언') or ''); ps = str(gvx(r, ix, '포지션') or '')
        if ch: t['champs'][ch] += 1
        if ps and ps != '선택안함': t['pos'][ps] += 1
        ai = None
        try: ai = float(gvx(r, ix, '점수'))
        except Exception: pass
        if ai is not None: t['aiS'] += ai; t['aiN'] += 1
        if ps and ps != '선택안함':
            pe = e['posst'][ps]; pe['g'] += 1; pe['w'] += (res == '승리')
            if ai is not None: pe['aiS'] += ai; pe['aiN'] += 1
        gid = str(gvx(r, ix, '게임ID') or '')
        if gid and '지표' in ix: games[gid].append((nm, res == '승리'))
        if P0 <= d <= P1:
            e['pw']['g'] += 1; e['pw']['w'] += (res == '승리')
        if W0 <= d <= W1:
            mt = metrics(gvx(r, ix, '지표'))
            g = {'d': str(gvx(r, ix, '날짜') or '')[5:16], 'champ': ch, 'pos': ps,
                 'kda': kda_str(gvx(r, ix, 'KDA')), 'res': res, 'eval': ev or None}
            if ai is not None: g['ai'] = round(ai, 1)
            if mt:
                m = mt['m']; dl = None
                try: dl = float(gvx(r, ix, '딜량'))
                except Exception: pass
                if dl and m: g['dpm'] = round(dl / m)
                if mt.get('cs') is not None and m: g['cspm'] = round(mt['cs'] / m, 1)
                if mt.get('kp') is not None: g['킬관여'] = int(mt['kp'])
                if mt.get('vs') is not None: g['시야'] = int(mt['vs'])
                if mt.get('sk'): g['솔킬'] = int(mt['sk'])
            e['tw'].append(g)

    # ── 🍀팀운 / ⚖️짊어진무게 (웹 v82.87 산식) ──
    career = {nm: e['tot']['aiS'] / e['tot']['aiN'] for nm, e in P.items() if e['tot']['aiN'] >= 5}
    acc = defaultdict(lambda: [0.0, 0.0, 0])   # raw합, implied합, n
    for gid, ps in games.items():
        if len(ps) < 8: continue
        for i, (me, win) in enumerate(ps):
            ms = mn = fs = fn = 0
            for j, (o, ow) in enumerate(ps):
                if j == i: continue
                c = career.get(o)
                if c is None: continue
                if ow == win: ms += c; mn += 1
                else: fs += c; fn += 1
            if mn < 3 or fn < 3: continue
            a = acc[me]
            a[0] += ms / mn - fs / fn; a[1] += 5 * (fs / fn) - 4 * (ms / mn); a[2] += 1
    elig = [(nm, s / n, t / n, n) for nm, (s, t, n) in acc.items() if n >= 10]
    LUCK, BURDEN = {}, {}
    if len(elig) >= 5:
        sx = sum(e[2] for e in elig); sy = sum(e[1] for e in elig); n = len(elig)
        sxx = sum(e[2] ** 2 for e in elig); sxy = sum(e[2] * e[1] for e in elig)
        b = (sxy - sx * sy / n) / max(1e-9, sxx - sx * sx / n); a0 = sy / n - b * (sx / n)
        resid = sorted((e[1] - (a0 + b * e[2]), e[0]) for e in elig)
        imps = sorted((e[2], e[0]) for e in elig)
        for i, (_, nm) in enumerate(resid): LUCK[nm] = round(i / (len(resid) - 1) * 100)
        for i, (_, nm) in enumerate(imps): BURDEN[nm] = round(i / (len(imps) - 1) * 100)

    # ── 🔀부포지션 (웹 산식) ──
    coh = defaultdict(lambda: {'wr': [], 'ai': []})
    for nm, e in P.items():
        for ps, st in e['posst'].items():
            if st['g'] >= 5:
                coh[ps]['wr'].append(st['w'] / st['g'])
                if st['aiN']: coh[ps]['ai'].append(st['aiS'] / st['aiN'])
    for ps in coh: coh[ps]['wr'].sort(); coh[ps]['ai'].sort()
    def pct(arr, v):
        if not arr or len(arr) < 5 or v is None: return None
        return round(bisect.bisect_left(arr, v) / len(arr) * 100)
    OFFROLE = {}
    for nm, e in P.items():
        if len(e['posst']) < 2: continue
        main = max(e['posst'], key=lambda k: e['posst'][k]['g'])
        ws = wn = 0; det = []
        for ps, st in e['posst'].items():
            if ps == main or st['g'] < 3: continue
            vals = [x for x in (pct(coh[ps]['wr'], st['w'] / st['g']),
                                pct(coh[ps]['ai'], st['aiS'] / st['aiN'] if st['aiN'] else None)) if x is not None]
            if not vals: continue
            ws += sum(vals) / len(vals) * st['g']; wn += st['g']; det.append(f"{ps}{st['g']}판")
        if wn: OFFROLE[nm] = {'score': round(ws / wn), 'detail': '·'.join(det), 'main': main}

    # ── 😤억울지수 (GRUDGE 탭) ──
    GR = {}
    if 'GRUDGE' in wb.sheetnames:
        rows = list(wb['GRUDGE'].values)
        if rows and rows[0] and '닉네임' in rows[0]:
            gi = {n: i for i, n in enumerate(rows[0])}
            for r in rows[1:]:
                try: GR[str(r[gi['닉네임']])] = float(r[gi['억울지수']])
                except Exception: pass

    out = {}
    for nm, e in P.items():
        if e['tot']['g'] < 10: continue
        t = e['tot']
        top = sorted(t['champs'].items(), key=lambda x: -x[1])[:6]
        mp = sorted(t['pos'].items(), key=lambda x: -x[1])[:2]
        tw = e['tw']; tww = sum(1 for g in tw if g['res'] == '승리')
        extra = {}
        if nm in LUCK: extra['팀운'] = LUCK[nm]
        if nm in BURDEN: extra['짊어진무게'] = BURDEN[nm]
        if nm in OFFROLE: extra['부포지션'] = OFFROLE[nm]
        if nm in GR: extra['억울지수'] = GR[nm]
        out[nm] = {'통산': {'판': t['g'], '승률': round(t['w'] / t['g'] * 100), 'MVP': t['mvp'],
                          'ACE': t['ace'], '역적': t['bad'],
                          '주챔프': [f"{c}{n}판" for c, n in top], '포지션': [f"{p}{n}" for p, n in mp]},
                   f'이번주({W0[5:].replace("-",".")}-{W1[5:].replace("-",".")})':
                       ({'판': len(tw), '승': tww, 'games': tw} if tw else None),
                   f'전주({P0[5:].replace("-",".")}-{P1[5:].replace("-",".")})':
                       (f"{e['pw']['g']}판 {round(e['pw']['w']/e['pw']['g']*100)}%" if e['pw']['g'] else '기록 없음'),
                   '신규지표': (extra or None),
                   '직전평가': prev_eval.get(nm, '')}
    return out

if __name__ == '__main__':
    mon = sys.argv[1]
    out = build(mon)
    path = f'digest_{mon.replace("-","")}.json'
    json.dump(out, open(path, 'w'), ensure_ascii=False)
    played = sum(1 for v in out.values() if any(k.startswith('이번주') and v[k] for k in v))
    print(f'{path} — {len(out)}명 (출전 {played})')
