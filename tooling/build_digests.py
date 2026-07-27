# AI 주간평 다이제스트 생성기 — squad_sheet.xlsx → digests/batch_NN.json
# 사용: python3 build_digests.py <squad_sheet.xlsx> <출력디렉토리> <이번주월요일 YYYY-MM-DD>
import sys, os, re, json, collections
import openpyxl
from datetime import date, timedelta

XLSX, OUT, WEEK = sys.argv[1], sys.argv[2], sys.argv[3]
wk = date.fromisoformat(WEEK)
PREV = (wk - timedelta(days=7)).isoformat()
WEEK_END = (wk + timedelta(days=7)).isoformat()

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
d = {n: list(wb[n].iter_rows(values_only=True)) for n in wb.sheetnames}
tn = lambda s: re.sub(r'\s+', '', str(s or '')).lower()

alt = {}
for r in d.get('LINK_ACCOUNT', [None])[1:]:
    if r and r[0] and r[1]: alt[tn(r[1])] = str(r[0]).strip()
canon = lambda n: alt.get(tn(n), str(n or '').strip())

def parse_metrics(s):
    o = {}
    for t in str(s or '').split('|'):
        m = re.match(r'^([a-z]+)(-?\d+(?:\.\d+)?)$', t.strip())
        if m: o[m.group(1)] = float(m.group(2))
    return o if o.get('m', 0) > 0 else None

P = collections.defaultdict(list)
for r in d['CLASSIC_NORMAL'][1:]:
    if r[8] in ('승리', '패배'): P[canon(r[2])].append(r)

sp = {tn(canon(r[1])): r for r in d['STAT_PLAYER'][1:]}
col = lambda tab, i=0: {tn(canon(r[i])): r for r in d.get(tab, [None])[1:] if r}
solo, peak, pos = col('SOLO_RANK'), col('PEAK_SEASONS'), col('CLAN_POSITIONS')
mbti, mito = col('MBTI'), col('MITO')
prev_eval = {tn(r[0]): r for r in d['AI_EVAL'][1:]}
champ = collections.defaultdict(list)
for r in d['STAT_CHAMP'][1:]:
    champ[tn(canon(r[1]))].append(r)

wr = lambda w, g: f"{round(100*w/g)}%" if g else "-"
digests = {}
for name, g in P.items():
    if len(g) < 10: continue
    k = tn(name); s = sp.get(k)
    wins = sum(1 for r in g if r[8] == '승리')
    pc, pw = collections.Counter(), collections.Counter()
    for r in g: pc[r[5]] += 1; pw[r[5]] += (r[8] == '승리')
    ch = sorted([r for r in champ.get(k, []) if r[4] and r[4] >= 3], key=lambda r: -r[4])[:8]
    def weeksum(start, end):
        rows = [r for r in g if r[1] and start <= str(r[1]) < end]
        if not rows: return None
        w = sum(1 for r in rows if r[8] == '승리')
        ev = collections.Counter(str(r[9]) for r in rows if str(r[9]) in ('MVP', '역적', 'ACE'))
        games = []
        for r in rows:
            mt = parse_metrics(r[18])
            it = {'날짜': str(r[1])[5:10], '포지션': r[5], '챔피언': r[6], '결과': r[8][0],
                  '평가': str(r[9]) if str(r[9]) in ('MVP', '역적', 'ACE') else '',
                  'KDA': r[11] or '', 'AI점수': r[12] or ''}
            if mt:
                dl = str(r[13] or '').replace(',', '')
                if dl.replace('.', '').isdigit(): it['분당딜'] = round(float(dl) / mt['m'])
                it['분당CS'] = round(mt.get('cs', 0) / mt['m'], 1)
                it['킬관여'] = int(mt.get('kp', 0)); it['시야'] = int(mt.get('vs', 0)); it['솔킬'] = int(mt.get('sk', 0))
            games.append(it)
        return {'판수': len(rows), '승': w, '승률': wr(w, len(rows)), '평가': dict(ev), '게임목록': games}
    this_w = weeksum(WEEK, WEEK_END)
    prev_w = weeksum(PREV, WEEK)
    digests[name] = {
        '닉네임': name,
        '통산': {'판수': len(g), '승률': wr(wins, len(g)),
                 'MVP': int(s[4]) if s and s[4] else 0, '역적': int(s[5]) if s and s[5] else 0,
                 'ACE': int(s[6]) if s and s[6] else 0,
                 '평균AI점수': round(s[7] / s[8], 1) if s and s[8] else None},
        '포지션별': {p: f"{c}판 {wr(pw[p], c)}" for p, c in pc.most_common() if p and p != '선택안함'},
        '주챔피언': {r[2]: f"{int(r[4])}판 {wr(r[5] or 0, r[4])}" + (f" ({r[3]})" if r[3] else "") for r in ch},
        '솔랭': (str(solo[k][1]) + ' ' + str(solo[k][2] or '') + 'LP' if k in solo else None),
        '최고티어': (f"{peak[k][2]} ({peak[k][3]})" if k in peak else None),
        '선언포지션': (f"{pos[k][2]}/{pos[k][3] or ''}" if k in pos else None),
        'MBTI': (str(mbti[k][1]) if k in mbti else None),
        '내전토너우승': (int(mito[k][1]) if k in mito and mito[k][1] else 0),
        f'이번주({WEEK[5:]}~)': this_w,
        f'전주({PREV[5:]}~)': ({'판수': prev_w['판수'], '승률': prev_w['승률'], '평가': prev_w['평가']} if prev_w else None),
        '직전평가': (str(prev_eval[k][1]) if k in prev_eval else None),
    }

os.makedirs(OUT, exist_ok=True)
names = sorted(digests, key=lambda n: -(digests[n][f'이번주({WEEK[5:]}~)'] or {'판수': 0})['판수'])
B = 10
for i in range(0, len(names), B):
    json.dump({n: digests[n] for n in names[i:i+B]},
              open(f'{OUT}/batch_{i//B:02d}.json', 'w'), ensure_ascii=False, indent=1)
json.dump(names, open(f'{OUT}/names.json', 'w'), ensure_ascii=False)
wkkey = f'이번주({WEEK[5:]}~)'
print(f'players={len(names)} batches={(len(names)+B-1)//B} active={sum(1 for n in names if digests[n][wkkey])}')
